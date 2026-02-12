"""
Phase 1: ICP Filter + Deduplication
Reads all 11 target lists, normalizes, filters by ICP, deduplicates, merges.
"""
import pandas as pd
import re
import os
import json
from rapidfuzz import fuzz, process
from collections import defaultdict

BASE = r"C:\Users\mamowi\Clients\AI-KIP\Target Lists unconsolidated"
OUT = r"C:\Users\mamowi\Clients\AI-KIP"

# ── ICP RULES ──────────────────────────────────────────────────────────────
AVOID_INDUSTRIES = {
    "banking", "insurance", "investment management", "financial services",
    "management consulting", "legal services", "accounting",
    "professional training & coaching", "staffing and recruiting",
    "retail", "consumer goods", "food & beverages", "leisure",
    "travel & tourism", "music", "education management",
    "hospital & health care", "apparel & fashion",
    "government administration", "public safety",
}

TIER1_INDUSTRIES = {
    "manufacturing", "industrial", "machinery", "mechanical or industrial engineering",
    "automotive", "chemicals", "building materials", "electrical/electronic manufacturing",
    "semiconductors", "plastics", "packaging", "metal", "textiles",
    "computer software", "information technology & services", "computer & network security",
    "computer hardware", "telecommunications", "internet",
    "renewables & environment", "defense & space",
}

TIER2_INDUSTRIES = {
    "wholesale", "distribution", "logistics & supply chain",
    "logistics and supply chain",
}

# ── HELPERS ────────────────────────────────────────────────────────────────

def parse_revenue_eur(val):
    """Normalize revenue to a single EUR number (float or None)."""
    if val is None or pd.isna(val):
        return None
    s = str(val).strip().replace("\u20ac", "").replace("€", "").replace(",", "")
    s = s.replace("~", "").replace(">", "").replace("<", "").replace("*", "")
    s = s.replace("bis ", "").replace("approx.", "").replace("approx", "")
    s = s.strip()
    if not s or s == "--" or s == "-":
        return None
    # Handle "10-50 Mio." style
    m = re.match(r"(\d+\.?\d*)\s*[-–]\s*(\d+\.?\d*)\s*(Mio\.?|M|Mrd\.?|B)", s, re.IGNORECASE)
    if m:
        lo, hi = float(m.group(1)), float(m.group(2))
        mid = (lo + hi) / 2
        unit = m.group(3).lower().rstrip(".")
        if unit in ("mio", "m"):
            return mid * 1_000_000
        elif unit in ("mrd", "b"):
            return mid * 1_000_000_000
        return mid
    # Handle "50 Mio. €" or "50M" or "479M"
    m = re.match(r"(\d+\.?\d*)\s*(Mio\.?|M|Mrd\.?|B)\b", s, re.IGNORECASE)
    if m:
        num = float(m.group(1))
        unit = m.group(2).lower().rstrip(".")
        if unit in ("mio", "m"):
            return num * 1_000_000
        elif unit in ("mrd", "b"):
            return num * 1_000_000_000
        return num
    # Handle "1.3B" or "1B"
    m = re.match(r"(\d+\.?\d*)\s*B$", s, re.IGNORECASE)
    if m:
        return float(m.group(1)) * 1_000_000_000
    # Handle plain large numbers (like 65000000)
    try:
        v = float(s)
        if v > 1000:  # likely raw EUR
            return v
        elif v > 0:
            return v * 1_000_000  # assume millions
        return None
    except ValueError:
        pass
    # Handle "XX Mio. €"
    m = re.match(r"(\d+\.?\d*)\s*Mio", s, re.IGNORECASE)
    if m:
        return float(m.group(1)) * 1_000_000
    return None


def parse_employees(val):
    """Normalize employee count to a single int or None."""
    if val is None or pd.isna(val):
        return None
    s = str(val).strip().replace("~", "").replace("+", "").replace(",", "")
    s = s.replace("est.", "").replace("est", "").replace("(", "").replace(")", "")
    s = s.replace("approx.", "").replace("approx", "").strip()
    if not s or s.lower() in ("mittelgroß", "familienunternehmen", "mittelgross",
                               "wachsend", "bundesweit", "international",
                               "division", "teil der mussmann-gruppe"):
        return None
    # Handle ranges like "200-499" or "20-49"
    m = re.match(r"(\d+)\s*[-–]\s*(\d+)", s)
    if m:
        return (int(m.group(1)) + int(m.group(2))) // 2
    # Handle employee class codes like "B 5000-10000" or "C 1000-4999"
    m = re.match(r"[A-Z]\s+(\d+)\s*-\s*(\d+)", s)
    if m:
        return (int(m.group(1)) + int(m.group(2))) // 2
    # Handle "150 Techniker" or "230 (CH)" etc - just grab first number
    m = re.match(r"(\d+)", s)
    if m:
        return int(m.group(1))
    return None


def classify_industry(industry_str):
    """Return ('tier1'|'tier2'|'avoid'|'unknown', cleaned_industry)."""
    if not industry_str or pd.isna(industry_str):
        return "unknown", ""
    s = str(industry_str).lower().strip()
    for avoid in AVOID_INDUSTRIES:
        if avoid in s:
            return "avoid", str(industry_str).strip()
    for t1 in TIER1_INDUSTRIES:
        if t1 in s:
            return "tier1", str(industry_str).strip()
    for t2 in TIER2_INDUSTRIES:
        if t2 in s:
            return "tier2", str(industry_str).strip()
    # Check German industry terms
    german_tier1 = ["messtechnik", "medizintechnik", "werkzeugbau", "automation",
                    "antriebstechnik", "it-services", "sicherheitstechnik",
                    "feinmechanik", "hydraulik", "pharmazeutik", "sondermaschinen",
                    "textiltechnik", "metallbau", "elektronik", "maschinenbau",
                    "automobilzulieferer", "industrie", "software", "it ",
                    "wasserspender", "kaffeemaschinen", "brandschutz", "arbeitsschutz",
                    "textilservice", "waschraumhygiene", "mattenservice", "led",
                    "luftreiniger", "büromöbel", "vending", "managed print",
                    "schädlingsbekämpfung", "büropflanzen", "erste hilfe",
                    "sicherheit", "kaffee", "wasser"]
    for term in german_tier1:
        if term in s:
            return "tier1", str(industry_str).strip()
    german_tier2 = ["großhandel", "distribution", "logistik",
                    "handel", "lebensmittelhandel"]
    for term in german_tier2:
        if term in s:
            return "tier2", str(industry_str).strip()
    german_avoid = ["beratung", "consulting", "rechtsanwalt", "versicherung",
                    "bank", "finanzen", "investment"]
    for term in german_avoid:
        if term in s:
            return "avoid", str(industry_str).strip()
    return "unknown", str(industry_str).strip()


def assign_client_tier(employees, revenue):
    """Assign client tier 1-5 based on ICP definition. Returns int or None."""
    tier_emp = None
    tier_rev = None
    if employees is not None:
        if employees >= 2000:
            tier_emp = 1
        elif employees >= 1000:
            tier_emp = 2
        elif employees >= 500:
            tier_emp = 3
        elif employees >= 200:
            tier_emp = 4
        elif employees >= 50:
            tier_emp = 5
        else:
            tier_emp = None  # too small
    if revenue is not None:
        if revenue >= 500_000_000:
            tier_rev = 1
        elif revenue >= 200_000_000:
            tier_rev = 2
        elif revenue >= 50_000_000:
            tier_rev = 3
        elif revenue >= 20_000_000:
            tier_rev = 4
        elif revenue >= 10_000_000:
            tier_rev = 5
        else:
            tier_rev = None  # below minimum
    # Use the higher tier (lower number = bigger company = higher tier)
    if tier_emp is not None and tier_rev is not None:
        return min(tier_emp, tier_rev)
    return tier_emp or tier_rev


def clean_company_name(name):
    """Normalize company name for matching."""
    if not name or pd.isna(name):
        return ""
    s = str(name).strip()
    # Remove common suffixes for matching purposes
    for suffix in [" GmbH & Co. KG", " GmbH & Co.KG", " GmbH & Co. KGaA",
                   " SE & Co. KGaA", " SE & Co. KG", " GmbH & Co.",
                   " GmbH", " AG", " SE", " KG", " e.V.", " OHG",
                   " mbH", " Co.", " Inc.", " Ltd."]:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
    return s


def normalize_country(val):
    """Normalize country to DE/AT/CH."""
    if not val or pd.isna(val):
        return ""
    s = str(val).strip().upper()
    mappings = {
        "DE": "DE", "GERMANY": "DE", "DEUTSCHLAND": "DE",
        "AT": "AT", "AUSTRIA": "AT", "ÖSTERREICH": "AT",
        "CH": "CH", "SWITZERLAND": "CH", "SCHWEIZ": "CH",
    }
    return mappings.get(s, s)


# ── READERS ────────────────────────────────────────────────────────────────

def read_prospecting_100():
    """List 1: AI-KIP_Prospecting_100_DACH_Companies.csv"""
    fp = os.path.join(BASE, "AI-KIP_Prospecting_100_DACH_Companies.csv")
    df = pd.read_csv(fp, encoding="utf-8")
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "company_name": str(r.get("Unternehmen", "")).strip(),
            "city": str(r.get("Stadt", "")).strip(),
            "country": normalize_country(r.get("Land")),
            "industry": str(r.get("Kategorie", "")).strip(),
            "description": str(r.get("Produkt", "")).strip(),
            "revenue_raw": r.get("Umsatz_Schaetzung_EUR"),
            "employees_raw": r.get("Mitarbeiter"),
            "domain": str(r.get("Website", "")).strip() if pd.notna(r.get("Website")) else "",
            "contact_name": str(r.get("Kontaktperson", "")).strip() if pd.notna(r.get("Kontaktperson")) else "",
            "contact_position": str(r.get("Position", "")).strip() if pd.notna(r.get("Position")) else "",
            "contact_linkedin": str(r.get("LinkedIn_URL", "")).strip() if pd.notna(r.get("LinkedIn_URL")) else "",
            "source": "AI-KIP_Prospecting_100",
        })
    return rows

def read_dach_10_50m():
    """List 2: dach_10_50m_targets.csv"""
    fp = os.path.join(BASE, "dach_10_50m_targets.csv")
    df = pd.read_csv(fp, encoding="utf-8")
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "company_name": str(r.get("Company", "")).strip(),
            "city": str(r.get("HQ City", "")).strip(),
            "country": normalize_country(r.get("Country")),
            "region": str(r.get("Region", "")).strip() if pd.notna(r.get("Region")) else "",
            "industry": str(r.get("Category", "")).strip() if pd.notna(r.get("Category")) else "",
            "description": str(r.get("What They Make / Sell", "")).strip() if pd.notna(r.get("What They Make / Sell")) else "",
            "revenue_raw": r.get("Approx Revenue EUR"),
            "employees_raw": r.get("Approx Employees"),
            "niche_leadership": str(r.get("Niche Leadership", "")).strip() if pd.notna(r.get("Niche Leadership")) else "",
            "source": "dach_10_50m_targets",
        })
    return rows

def read_dach_icp():
    """List 3: dach_icp_targets.csv - has commas inside bracket fields, need quotechar handling"""
    fp = os.path.join(BASE, "dach_icp_targets.csv")
    # Read raw lines and rejoin fields that were split inside brackets
    import csv
    with open(fp, "r", encoding="utf-8") as f:
        reader = csv.reader(f, quotechar='"')
        header = next(reader)
        raw_rows = []
        for line in reader:
            # If line has more fields than header, merge extra fields into Sub-Industry
            if len(line) > len(header):
                # Fields 7 onwards might be split Sub-Industry - rejoin
                diff = len(line) - len(header)
                fixed = line[:7] + [",".join(line[7:7+diff+1])] + line[7+diff+1:]
                raw_rows.append(dict(zip(header, fixed)))
            else:
                raw_rows.append(dict(zip(header, line)))
    df = pd.DataFrame(raw_rows)
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "company_name": str(r.get("Company", "")).strip().strip('"'),
            "city": str(r.get("City", "")).strip().strip('"'),
            "country": normalize_country(r.get("Country")),
            "industry": str(r.get("Industry", "")).strip().strip('"'),
            "sub_industry": str(r.get("Sub-Industry", "")).strip().strip('"') if pd.notna(r.get("Sub-Industry")) else "",
            "revenue_raw": r.get("Revenue (EUR)"),
            "employees_raw": r.get("Employees"),
            "domain": str(r.get("Domain", "")).strip().strip('"') if pd.notna(r.get("Domain")) else "",
            "icp_score": r.get("ICP Score"),
            "sales_team": r.get("Sales Team"),
            "offices": r.get("Offices"),
            "icp_reasons": str(r.get("ICP Reasons", "")).strip().strip('"') if pd.notna(r.get("ICP Reasons")) else "",
            "source": "dach_icp_targets",
        })
    return rows

def read_mittelstand_manufacturers():
    """List 4: dach_mittelstand_manufacturers_to_source.csv"""
    fp = os.path.join(BASE, "dach_mittelstand_manufacturers_to_source.csv")
    df = pd.read_csv(fp, encoding="utf-8")
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "company_name": str(r.get("Company", "")).strip(),
            "city": str(r.get("HQ City", "")).strip(),
            "region": str(r.get("State/Region", "")).strip() if pd.notna(r.get("State/Region")) else "",
            "country": normalize_country(r.get("Country")),
            "industry": "Manufacturing",
            "description": str(r.get("What They Make / Niche", "")).strip() if pd.notna(r.get("What They Make / Niche")) else "",
            "revenue_raw": r.get("Approx Revenue (EUR)"),
            "employees_raw": r.get("Approx Employees"),
            "world_market_leader": str(r.get("World Market Leader?", "")).strip() if pd.notna(r.get("World Market Leader?")) else "",
            "priority": str(r.get("Priority", "")).strip() if pd.notna(r.get("Priority")) else "",
            "source": "dach_mittelstand_manufacturers",
        })
    return rows

def read_si_impact():
    """List 5: dach_si_impact_top100.csv"""
    fp = os.path.join(BASE, "dach_si_impact_top100.csv")
    df = pd.read_csv(fp, encoding="utf-8")
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "company_name": str(r.get("Company", "")).strip().strip('"'),
            "city": str(r.get("City", "")).strip().strip('"'),
            "country": normalize_country(r.get("Country")),
            "industry": str(r.get("Industry", "")).strip().strip('"'),
            "revenue_raw": r.get("Revenue"),
            "employees_raw": r.get("Employees"),
            "domain": str(r.get("Domain", "")).strip().strip('"') if pd.notna(r.get("Domain")) else "",
            "impact_score": r.get("Impact Score"),
            "sales_team": r.get("Sales"),
            "offices": r.get("Offices"),
            "source": "dach_si_impact_top100",
        })
    return rows

def read_zollernalbkreis_complete():
    """List 7: Zollernalbkreis_Sales_Contacts_Complete.csv (superset of list 6)"""
    fp = os.path.join(BASE, "Zollernalbkreis_Sales_Contacts_Complete.csv")
    df = pd.read_csv(fp, encoding="utf-8")
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "company_name": str(r.get("Unternehmen", "")).strip(),
            "city": str(r.get("Standort", "")).strip(),
            "country": "DE",
            "industry": str(r.get("Branche", "")).strip() if pd.notna(r.get("Branche")) else "",
            "revenue_raw": r.get("Umsatz"),
            "domain": str(r.get("Domain", "")).strip() if pd.notna(r.get("Domain")) else "",
            "contact_name": str(r.get("Kontaktperson", "")).strip() if pd.notna(r.get("Kontaktperson")) else "",
            "contact_position": str(r.get("Position", "")).strip() if pd.notna(r.get("Position")) else "",
            "contact_linkedin": str(r.get("LinkedIn URL", "")).strip() if pd.notna(r.get("LinkedIn URL")) else "",
            "contact_email": str(r.get("Email (predicted)", "")).strip() if pd.notna(r.get("Email (predicted)")) else "",
            "source": "Zollernalbkreis_Contacts",
        })
    return rows

def read_familienunternehmen_v6():
    """List 11: Top Familienunternehmen DACH V6 100Mio-1Mrd.xlsx"""
    import openpyxl
    fp = os.path.join(BASE, "Top Familienunternehmen DACH V6 100Mio-1Mrd.xlsx")
    wb = openpyxl.load_workbook(fp, read_only=True)
    ws = wb["DACH FamU V6"]
    rows_out = []
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
            continue
        d = dict(zip(header, row))
        company = str(d.get("Unternehmen", "")).strip()
        if not company:
            continue
        country = normalize_country(d.get("Land", ""))
        # Parse GF names
        gf_names = []
        for idx in range(1, 6):
            fn = d.get(f"Vorname {idx}. GF", "")
            ln = d.get(f"Name {idx}. GF", "")
            title = d.get(f"Titel {idx}. GF", "")
            salut = d.get(f"Anrede {idx}. GF/CEO", "") if idx == 1 else d.get(f"Anrede {idx}. GF/CEO", "")
            if fn and ln and not pd.isna(fn) and not pd.isna(ln):
                full = f"{str(title).strip()} {str(fn).strip()} {str(ln).strip()}".strip()
                gf_names.append(full)
        rows_out.append({
            "company_name": company,
            "company_name_2": str(d.get("Unternehmen 2", "")).strip() if d.get("Unternehmen 2") and not pd.isna(d.get("Unternehmen 2")) else "",
            "city": str(d.get("Ort", "")).strip() if d.get("Ort") else "",
            "zip": str(d.get("PLZ", "")).strip() if d.get("PLZ") else "",
            "region": str(d.get("Bundesland / Kanton", "")).strip() if d.get("Bundesland / Kanton") else "",
            "country": country,
            "phone": str(d.get("Telefon", "")).strip() if d.get("Telefon") and not pd.isna(d.get("Telefon")) else "",
            "domain": str(d.get("Website", "")).strip().replace("http://", "").replace("https://", "").rstrip("/") if d.get("Website") and not pd.isna(d.get("Website")) else "",
            "email": str(d.get("Mailadresse", "")).strip() if d.get("Mailadresse") and not pd.isna(d.get("Mailadresse")) else "",
            "industry": str(d.get("Branche", "")).strip() if d.get("Branche") and not pd.isna(d.get("Branche")) else "",
            "industry_category": str(d.get("Oberkategorie", "")).strip() if d.get("Oberkategorie") and not pd.isna(d.get("Oberkategorie")) else "",
            "segment": str(d.get("Segment", "")).strip() if d.get("Segment") and not pd.isna(d.get("Segment")) else "",
            "description": str(d.get("Kurzbeschreibung", "")).strip() if d.get("Kurzbeschreibung") and not pd.isna(d.get("Kurzbeschreibung")) else "",
            "founding_year": d.get("Gründungs-Jahr") if d.get("Gründungs-Jahr") and not pd.isna(d.get("Gründungs-Jahr")) else None,
            "ownership": str(d.get("Familie/n / Investor / Eigentümer", "")).strip() if d.get("Familie/n / Investor / Eigentümer") and not pd.isna(d.get("Familie/n / Investor / Eigentümer")) else "",
            "revenue_raw": d.get("Umsatz in Mio. Euro"),
            "employees_raw": d.get("Mitarbeiter"),
            "geschaeftsfuehrer": gf_names,
            "source": "Familienunternehmen_V6",
        })
    wb.close()
    return rows_out

def read_research_xlsx():
    """List 10: Target list for additional Research.xlsx"""
    import openpyxl
    fp = os.path.join(BASE, "Target list for additional Reserach.xlsx")
    wb = openpyxl.load_workbook(fp, read_only=True)
    ws = wb["Sheet1"]
    rows_out = []
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
            continue
        d = dict(zip(header, row))
        company = str(d.get("Unternehmen", "")).strip()
        if not company:
            continue
        rows_out.append({
            "company_name": company,
            "city": str(d.get("Ort", "")).strip() if d.get("Ort") else "",
            "zip": str(d.get("PLZ", "")).strip() if d.get("PLZ") else "",
            "b2b_flag": str(d.get("Geschäftsausrichtung", "")).strip() if d.get("Geschäftsausrichtung") and not pd.isna(d.get("Geschäftsausrichtung")) else "",
            "linkedin_company": str(d.get("LinkedIN Profile URL", "")).strip() if d.get("LinkedIN Profile URL") and not pd.isna(d.get("LinkedIN Profile URL")) else "",
            "source": "Research_xlsx",
        })
    wb.close()
    return rows_out

def parse_md_manufacturers():
    """List 8: DACH_Mittelstand_B2B_Manufacturers_ICP_List.md"""
    fp = os.path.join(BASE, "DACH_Mittelstand_B2B_Manufacturers_ICP_List.md")
    with open(fp, "r", encoding="utf-8") as f:
        content = f.read()
    rows = []
    # Parse ### blocks
    blocks = re.split(r"###\s+\d+\.\s+", content)
    for block in blocks[1:]:  # skip preamble
        lines = block.strip().split("\n")
        name = lines[0].strip()
        data = {}
        for line in lines:
            m = re.match(r"\|\s*\*\*(.+?)\*\*\s*\|\s*(.+?)\s*\|", line)
            if m:
                data[m.group(1).strip()] = m.group(2).strip()
        hq = data.get("HQ", "")
        city = ""
        region = ""
        country = "DE"
        if "Austria" in hq or "Österreich" in hq:
            country = "AT"
        elif "Switzerland" in hq or "Schweiz" in hq:
            country = "CH"
        city_parts = hq.split(",")
        if city_parts:
            city = city_parts[0].strip()
        if len(city_parts) > 1:
            region = city_parts[1].strip()
        rows.append({
            "company_name": name,
            "city": city,
            "region": region,
            "country": country,
            "industry": "Manufacturing",
            "description": data.get("What they make", ""),
            "revenue_raw": data.get("Revenue", ""),
            "employees_raw": data.get("Employees", ""),
            "world_market_leader": "Yes" if "leader" in data.get("Why they fit", "").lower() or "weltmarkt" in data.get("Why they fit", "").lower() else "",
            "source": "MD_Manufacturers_ICP",
        })
    return rows

def parse_md_hidden_champions():
    """List 9: DACH_Small_Hidden_Champions_10-50M.md"""
    fp = os.path.join(BASE, "DACH_Small_Hidden_Champions_10-50M.md")
    with open(fp, "r", encoding="utf-8") as f:
        content = f.read()
    rows = []
    # Parse markdown tables
    table_pattern = re.compile(
        r"\|\s*(\d+)\s*\|\s*\*\*(.+?)\*\*\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*([\d~.]+.*?)\s*\|\s*([\d~.]+.*?)\s*\|\s*(.+?)\s*\|"
    )
    for m in table_pattern.finditer(content):
        company = m.group(2).strip()
        city = m.group(3).strip()
        region = m.group(4).strip()
        description = m.group(5).strip()
        employees = m.group(6).strip()
        revenue = m.group(7).strip()
        niche = m.group(8).strip()
        country = "DE"
        if "Austria" in region or "Österreich" in region or "AT" in region:
            country = "AT"
        elif "Switzerland" in region or "Schweiz" in region or "CH" in region:
            country = "CH"
        rows.append({
            "company_name": company,
            "city": city,
            "region": region,
            "country": country,
            "industry": "Manufacturing",
            "description": description,
            "revenue_raw": revenue,
            "employees_raw": employees,
            "niche_leadership": niche,
            "source": "MD_Hidden_Champions",
        })
    # Also parse tables without bold but with | # | Company | pattern (Tier 5 style)
    table2 = re.compile(
        r"\|\s*(\d+)\s*\|\s*\*\*(.+?)\*\*\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*([\d~.]+.*?)\s*\|\s*(.+?)\s*\|"
    )
    # Already captured above. Skip duplicates.
    return rows


# ── MAIN ───────────────────────────────────────────────────────────────────

print("=" * 70)
print("PHASE 1: ICP FILTER + DEDUPLICATION")
print("=" * 70)

# Step 1: Read all sources
print("\n[1/6] Reading all 11 source files...")

all_rows = []

readers = [
    ("Prospecting 100", read_prospecting_100),
    ("DACH 10-50M", read_dach_10_50m),
    ("DACH ICP Targets", read_dach_icp),
    ("Mittelstand Manufacturers", read_mittelstand_manufacturers),
    ("SI Impact Top 100", read_si_impact),
    ("Zollernalbkreis Contacts", read_zollernalbkreis_complete),
    ("MD Manufacturers ICP", parse_md_manufacturers),
    ("MD Hidden Champions", parse_md_hidden_champions),
    ("Research XLSX", read_research_xlsx),
    ("Familienunternehmen V6", read_familienunternehmen_v6),
]

for name, reader in readers:
    try:
        rows = reader()
        print(f"  {name}: {len(rows)} entries")
        all_rows.extend(rows)
    except Exception as e:
        print(f"  {name}: ERROR - {e}")

print(f"\n  TOTAL RAW ENTRIES: {len(all_rows)}")

# Step 2: Normalize revenue and employees
print("\n[2/6] Normalizing revenue & employee data...")
for row in all_rows:
    rev_raw = row.get("revenue_raw")
    # Special handling for V6 which has revenue in Mio already
    if row.get("source") == "Familienunternehmen_V6" and rev_raw is not None and not pd.isna(rev_raw):
        try:
            v = float(rev_raw)
            row["revenue_eur"] = v * 1_000_000  # V6 stores in Mio
        except (ValueError, TypeError):
            row["revenue_eur"] = parse_revenue_eur(rev_raw)
    else:
        row["revenue_eur"] = parse_revenue_eur(rev_raw)
    row["employees"] = parse_employees(row.get("employees_raw"))

# Step 3: ICP Filter
print("\n[3/6] Applying ICP filters...")
icp_eligible = []
filtered_out = defaultdict(int)

for row in all_rows:
    name = row.get("company_name", "").strip()
    if not name:
        filtered_out["empty_name"] += 1
        continue

    country = row.get("country", "")
    if country and country not in ("DE", "AT", "CH", ""):
        filtered_out["non_dach"] += 1
        continue

    employees = row.get("employees")
    if employees is not None and employees < 20:
        filtered_out["too_small"] += 1
        continue

    # Industry classification
    industry = row.get("industry", "")
    industry_cat = row.get("industry_category", "")
    segment = row.get("segment", "")
    # Try industry first, then category, then segment
    ind_tier, ind_clean = classify_industry(industry)
    if ind_tier == "unknown" and industry_cat:
        ind_tier, ind_clean = classify_industry(industry_cat)
    if ind_tier == "unknown" and segment:
        ind_tier, ind_clean = classify_industry(segment)
    # B2B flag from research xlsx
    b2b_flag = row.get("b2b_flag", "")
    if b2b_flag and "b2c" in b2b_flag.lower() and "b2b" not in b2b_flag.lower():
        filtered_out["b2c"] += 1
        continue

    if ind_tier == "avoid":
        filtered_out["avoid_industry"] += 1
        continue

    row["industry_tier"] = ind_tier
    row["industry_clean"] = ind_clean
    row["client_tier"] = assign_client_tier(employees, row.get("revenue_eur"))

    icp_eligible.append(row)

print(f"  ICP-eligible: {len(icp_eligible)}")
print(f"  Filtered out: {sum(filtered_out.values())}")
for reason, count in sorted(filtered_out.items(), key=lambda x: -x[1]):
    print(f"    {reason}: {count}")

# Step 4: Deduplication via fuzzy matching
print("\n[4/6] Deduplicating (fuzzy match on company name + city)...")

# Build match key
for row in icp_eligible:
    cn = clean_company_name(row.get("company_name", ""))
    city = str(row.get("city", "")).strip().lower()
    row["_match_key"] = f"{cn.lower()}|{city}"
    row["_clean_name"] = cn.lower()

# Group by exact match key first
exact_groups = defaultdict(list)
for row in icp_eligible:
    exact_groups[row["_match_key"]].append(row)

# Now fuzzy-match across groups
group_keys = list(exact_groups.keys())
merged_groups = []
used = set()

for i, key1 in enumerate(group_keys):
    if key1 in used:
        continue
    group = list(exact_groups[key1])
    used.add(key1)
    name1 = key1.split("|")[0]
    for j in range(i + 1, len(group_keys)):
        key2 = group_keys[j]
        if key2 in used:
            continue
        name2 = key2.split("|")[0]
        # Fuzzy match on name
        score = fuzz.token_sort_ratio(name1, name2)
        if score >= 85:
            group.extend(exact_groups[key2])
            used.add(key2)
    merged_groups.append(group)

print(f"  Unique companies after dedup: {len(merged_groups)}")

# Step 5: Merge data within each group (richest source wins)
print("\n[5/6] Merging data from multiple sources per company...")

SOURCE_PRIORITY = {
    "Familienunternehmen_V6": 10,
    "AI-KIP_Prospecting_100": 9,
    "Zollernalbkreis_Contacts": 8,
    "dach_icp_targets": 7,
    "dach_si_impact_top100": 6,
    "dach_mittelstand_manufacturers": 5,
    "dach_10_50m_targets": 4,
    "MD_Manufacturers_ICP": 3,
    "MD_Hidden_Champions": 2,
    "Research_xlsx": 1,
}

master = []

for group in merged_groups:
    # Sort by source priority (highest first)
    group.sort(key=lambda r: SOURCE_PRIORITY.get(r.get("source", ""), 0), reverse=True)
    merged = {}
    sources = set()
    contacts = []

    for row in group:
        sources.add(row.get("source", ""))
        # For each field, take the first non-empty value (from highest priority source)
        for field in ["company_name", "city", "region", "country", "zip", "phone",
                      "domain", "email", "industry", "industry_clean", "industry_category",
                      "segment", "description", "founding_year", "ownership",
                      "niche_leadership", "world_market_leader", "priority",
                      "icp_score", "impact_score", "icp_reasons",
                      "industry_tier", "b2b_flag", "linkedin_company",
                      "sales_team", "offices"]:
            if field not in merged or not merged[field]:
                val = row.get(field)
                if val and not pd.isna(val) and str(val).strip():
                    merged[field] = val

        # Revenue: take the most precise (non-None) value
        if "revenue_eur" not in merged or merged["revenue_eur"] is None:
            if row.get("revenue_eur") is not None:
                merged["revenue_eur"] = row["revenue_eur"]
                merged["revenue_raw"] = row.get("revenue_raw")

        # Employees: take the most precise value
        if "employees" not in merged or merged["employees"] is None:
            if row.get("employees") is not None:
                merged["employees"] = row["employees"]

        # Client tier: recalculate after merge
        # Contacts: collect all
        if row.get("contact_name"):
            contacts.append({
                "name": row["contact_name"],
                "position": row.get("contact_position", ""),
                "email": row.get("contact_email", ""),
                "linkedin": row.get("contact_linkedin", ""),
            })
        # GF from V6
        if row.get("geschaeftsfuehrer"):
            for gf in row["geschaeftsfuehrer"]:
                if gf and gf not in [c["name"] for c in contacts]:
                    contacts.append({
                        "name": gf,
                        "position": "Geschäftsführer",
                        "email": "",
                        "linkedin": "",
                    })

    # Recalculate tier
    merged["client_tier"] = assign_client_tier(merged.get("employees"), merged.get("revenue_eur"))
    merged["sources"] = "; ".join(sorted(sources))
    merged["source_count"] = len(sources)
    merged["contacts"] = contacts
    merged["has_url"] = bool(merged.get("domain"))
    merged["has_contact"] = len(contacts) > 0

    # Data completeness score
    completeness = 0
    if merged.get("domain"):
        completeness += 20
    if merged.get("revenue_eur"):
        completeness += 20
    if merged.get("employees"):
        completeness += 20
    if merged.get("industry_clean") or merged.get("industry"):
        completeness += 20
    if contacts:
        completeness += 20
    merged["data_completeness"] = completeness

    master.append(merged)

# Step 6: Sort and output
print("\n[6/6] Generating output...")

# Sort: by client_tier (ascending, best first), then data completeness (desc)
master.sort(key=lambda r: (
    r.get("client_tier") or 99,
    -r.get("data_completeness", 0),
    -(r.get("source_count", 0)),
))

# Build output DataFrame
out_rows = []
for m in master:
    primary_contact = m.get("contacts", [{}])[0] if m.get("contacts") else {}
    out_rows.append({
        "company_name": m.get("company_name", ""),
        "domain": m.get("domain", ""),
        "city": m.get("city", ""),
        "region": m.get("region", ""),
        "country": m.get("country", ""),
        "zip": m.get("zip", ""),
        "phone": m.get("phone", ""),
        "email_company": m.get("email", ""),
        "industry": m.get("industry_clean") or m.get("industry", ""),
        "industry_tier": m.get("industry_tier", ""),
        "description": m.get("description", ""),
        "revenue_eur": m.get("revenue_eur"),
        "employees": m.get("employees"),
        "client_tier": m.get("client_tier"),
        "ownership": m.get("ownership", ""),
        "founding_year": m.get("founding_year"),
        "world_market_leader": m.get("world_market_leader", ""),
        "niche_leadership": m.get("niche_leadership", ""),
        "contact_1_name": primary_contact.get("name", ""),
        "contact_1_position": primary_contact.get("position", ""),
        "contact_1_email": primary_contact.get("email", ""),
        "contact_1_linkedin": primary_contact.get("linkedin", ""),
        "has_url": m.get("has_url"),
        "has_contact": m.get("has_contact"),
        "data_completeness": m.get("data_completeness"),
        "sources": m.get("sources", ""),
        "source_count": m.get("source_count", 0),
    })

df_out = pd.DataFrame(out_rows)
outpath = os.path.join(OUT, "master_target_list_phase1.csv")
df_out.to_csv(outpath, index=False, encoding="utf-8-sig")
print(f"\n  Master list saved: {outpath}")

# ── STATISTICS ─────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("PHASE 1 RESULTS")
print("=" * 70)

total = len(master)
print(f"\nTotal unique ICP-eligible companies: {total}")

# Tier distribution
tier_counts = defaultdict(int)
for m in master:
    t = m.get("client_tier")
    tier_counts[t if t else "unassigned"] += 1
print("\nClient Tier Distribution:")
for t in sorted(tier_counts.keys(), key=lambda x: (isinstance(x, str), x)):
    label = f"Tier {t}" if isinstance(t, int) else t
    print(f"  {label}: {tier_counts[t]}")

# Industry tier
ind_counts = defaultdict(int)
for m in master:
    ind_counts[m.get("industry_tier", "unknown")] += 1
print("\nIndustry Tier Distribution:")
for t, c in sorted(ind_counts.items()):
    print(f"  {t}: {c}")

# Country
country_counts = defaultdict(int)
for m in master:
    country_counts[m.get("country", "unknown")] += 1
print("\nCountry Distribution:")
for c, n in sorted(country_counts.items(), key=lambda x: -x[1]):
    print(f"  {c}: {n}")

# Data gaps
has_url = sum(1 for m in master if m.get("has_url"))
has_contact = sum(1 for m in master if m.get("has_contact"))
has_revenue = sum(1 for m in master if m.get("revenue_eur"))
has_employees = sum(1 for m in master if m.get("employees"))
has_tier = sum(1 for m in master if m.get("client_tier"))

print("\nData Completeness:")
print(f"  Has URL:        {has_url:>5} / {total}  ({has_url*100//total}%)")
print(f"  Has Contact:    {has_contact:>5} / {total}  ({has_contact*100//total}%)")
print(f"  Has Revenue:    {has_revenue:>5} / {total}  ({has_revenue*100//total}%)")
print(f"  Has Employees:  {has_employees:>5} / {total}  ({has_employees*100//total}%)")
print(f"  Has Tier:       {has_tier:>5} / {total}  ({has_tier*100//total}%)")

# Multi-source companies (appeared in 2+ lists)
multi = sum(1 for m in master if m.get("source_count", 0) >= 2)
print(f"\nMulti-source companies (in 2+ lists): {multi}")

# Completeness distribution
comp_dist = defaultdict(int)
for m in master:
    comp_dist[m.get("data_completeness", 0)] += 1
print("\nData Completeness Score Distribution:")
for score in sorted(comp_dist.keys()):
    print(f"  {score}%: {comp_dist[score]} companies")

# What needs to happen next
no_url = total - has_url
no_contact = total - has_contact
no_tier = total - has_tier
print(f"\n--- GAPS TO FILL ---")
print(f"  Companies missing URL (MUST):    {no_url}")
print(f"  Companies missing Contact:       {no_contact}")
print(f"  Companies missing Tier (no rev/emp): {no_tier}")
